#!/usr/bin/env -S uv run --quiet --script
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
#
# @description Convert json/jsonl to a formatted xlsx spreadsheet
# @usage tiss json2xlsx <out.xlsx> [--sheet NAME]
# @example tiss csv2json data.csv | tiss json2xlsx report.xlsx
# @example tiss readData aws/params | tiss json2xlsx params.xlsx --sheet Params
# @needs uv
#
# The polyglot leaf: a python script in the tiss tree, self-contained via
# uv + PEP 723 inline dependencies — the dispatcher runs it like any other
# command. Reads jsonl (or a JSON array) on stdin and writes a spreadsheet
# with a bold frozen header, auto-sized columns, comma-formatted numbers,
# and real dates for ISO-looking strings.
#
import json
import re
import sys
from datetime import datetime, timezone

ISO_DT = re.compile(r"^\d{4}-\d{2}-\d{2}([T ]\d{2}:\d{2}(:\d{2})?(\.\d+)?(Z|[+-]\d{2}:?\d{2})?)?$")


def read_records(stream):
    """Accept jsonl (one object per line) or a single JSON array/object."""
    text = stream.read().strip()
    if not text:
        return []
    if text.startswith("["):
        data = json.loads(text)
        return data if isinstance(data, list) else [data]
    records = []
    for line in text.splitlines():
        line = line.strip()
        if line:
            records.append(json.loads(line))
    return records


def coerce(value):
    """Turn ISO-date strings into datetimes so excel treats them as dates."""
    if isinstance(value, str) and ISO_DT.match(value):
        try:
            dt = datetime.fromisoformat(value.replace("Z", "+00:00").replace(" ", "T"))
            if dt.tzinfo is not None:
                # Excel cannot store tz-aware datetimes: normalize to UTC, drop tz.
                dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
            return dt
        except ValueError:
            return value
    if isinstance(value, (dict, list)):
        return json.dumps(value)  # nested structures stay readable
    return value


def main():
    args = sys.argv[1:]
    if not args or args[0] in ("-h", "--help"):
        print("usage: tiss json2xlsx <out.xlsx> [--sheet NAME]", file=sys.stderr)
        sys.exit(0 if args else 2)

    out = args[0]
    sheet = "Data"
    if "--sheet" in args:
        sheet = args[args.index("--sheet") + 1]

    records = read_records(sys.stdin)
    if not records:
        print("json2xlsx: no records on stdin", file=sys.stderr)
        sys.exit(1)

    # Header: union of keys, in first-seen order.
    headers = []
    for rec in records:
        for key in rec:
            if key not in headers:
                headers.append(key)

    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    widths = [len(h) for h in headers]
    for rec in records:
        row = [coerce(rec.get(h)) for h in headers]
        ws.append(row)
        for i, v in enumerate(row):
            widths[i] = max(widths[i], len(str(v)) if v is not None else 0)

    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = min(width + 2, 60)
        for cell in ws[get_column_letter(col)][1:]:
            if isinstance(cell.value, float):
                cell.number_format = "#,##0.00"
            elif isinstance(cell.value, int):
                cell.number_format = "#,##0"
            elif isinstance(cell.value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"

    wb.save(out)
    print(f"json2xlsx: wrote {len(records)} rows -> {out}", file=sys.stderr)


if __name__ == "__main__":
    main()
